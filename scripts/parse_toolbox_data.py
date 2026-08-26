import sys
import os
import openpyxl
import json
import re

sys.stdout.reconfigure(encoding='utf-8')

tools_dir = r'C:\Users\Motasem.ghanem\Downloads\Techincians tools'

# Status colors
# Yellow: FFFFFF00 -> damaged
# Red: FFFF0000 -> missing
# Purple: FF7030A0 -> not_delivered

def get_cell_status(cell):
    if not cell or not cell.fill or not cell.fill.start_color:
        return 'good'
    sc = cell.fill.start_color
    rgb = str(sc.rgb) if sc.rgb else ''
    if 'FFFFFF00' in rgb or rgb == 'FFFF00':
        return 'damaged'
    if 'FFFF0000' in rgb or rgb == 'FF0000':
        return 'missing'
    if 'FF7030A0' in rgb or rgb == '7030A0':
        return 'not_delivered'
    if sc.theme == 5 and sc.tint is not None and abs(sc.tint - (-0.249977111117893)) < 0.01:
        return 'not_delivered'
    return 'good'

def clean_val(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None

def categorize(cat_title, item_name):
    cat = str(cat_title or '').strip()
    item = str(item_name or '').strip()
    
    if any(k in cat for k in ['بكسات', 'بوكس', 'اسود طويل']):
        return 'sockets', 'طقم بكسات وبوكسات', 'Socket Sets'
    if any(k in cat for k in ['شق-رنج', 'شق -رنج', 'شق_رنج']):
        return 'combination_wrenches', 'مفاتيح شق-رنج', 'Combination Spanners'
    if any(k in cat for k in ['شق-شق', 'شق -شق', 'مفتاح شق']):
        return 'open_wrenches', 'مفاتيح شق', 'Open-End Spanners'
    if any(k in cat for k in ['مشرشر', 'Torx', 'torx']):
        return 'torx_keys', 'مفاتيح ومشرشر Torx', 'Torx & Star Keys'
    if any(k in cat for k in ['النكيه', 'الينكيه', 'النيكيه', 'Hex']):
        return 'hex_keys', 'مفاتيح ألنكيه (Hex)', 'Allen / Hex Keys'
    if any(k in cat for k in ['مفكات', 'مفك']):
        return 'screwdrivers', 'مفكات متنوعة', 'Screwdrivers'
    if any(k in cat for k in ['سناب رنج']):
        return 'snap_rings', 'طقم سناب رنج', 'Snap Ring Pliers'
    if any(k in cat for k in ['مبرد']):
        return 'files', 'طقم مبارد', 'Files Set'
        
    if any(k in item for k in ['يدة', 'يد ', 'وصلة']):
        return 'ratchets_extensions', 'يدات توصيل ووصلات', 'Ratchets & Extensions'
    if any(k in item for k in ['زرادية', 'قطاعة', 'بنسة', 'كماشة', 'بيلاروس']):
        return 'pliers_cutters', 'زراديات وبنس وقواطع', 'Pliers & Cutters'
    if any(k in item for k in ['شاكوش', 'مطرقة', 'منشار', 'مبرد', 'مشحاف']):
        return 'hammers_saws', 'مطارق ومناشير ومبارد', 'Hammers & Saws'
    if any(k in item for k in ['ميتر', 'Meter', 'متر', 'قياس', 'كشاف', 'مصباح', 'مغناطيس']):
        return 'electrical_measuring', 'أجهزة قياس وكهرباء وإضاءة', 'Measurement & Electrical'
    if any(k in item for k in ['صندوق', 'شنتة', 'حقيبة']):
        return 'storage', 'صناديق وحقائب عدة', 'Toolboxes & Storage'
    if any(k in item for k in ['فرد شحم', 'مفتاح بايبات', 'مفتاح بسطون', 'مفتاح فلاتر', 'كتر', 'تحويلة', 'سنبك']):
        return 'specialty_tools', 'أدوات صيانة متخصصة', 'Specialty Maintenance Tools'
        
    return 'general_tools', 'أدوات عامة', 'General Tools'

def get_icon(category_id, item_name):
    item = str(item_name).lower()
    if 'multi meter' in item or 'meter' in item or 'ميتر' in item:
        return 'multimeter'
    if 'متر' in item or 'قياس' in item:
        return 'measuring_tape'
    if 'مصباح' in item or 'كشاف' in item:
        return 'flashlight'
    if 'مغناطيس' in item:
        return 'magnet'
    if 'شاكوش' in item or 'مطرقة' in item:
        return 'hammer'
    if 'منشار' in item:
        return 'saw'
    if 'مبرد' in item:
        return 'file_tool'
    if 'قطاعة' in item or 'تعرية' in item:
        return 'cutter'
    if 'زرادية' in item or 'بنسة' in item or 'كماشة' in item or 'بيلاروس' in item or 'سناب' in item:
        return 'pliers'
    if 'يدة' in item or 'اوتوماتك' in item or 'ratchet' in item:
        return 'ratchet'
    if 'وصلة' in item or 'تحويلة' in item:
        return 'extension_bar'
    if 'مفك مصلب' in item or 'philips' in item or 'cross' in item:
        return 'screwdriver_cross'
    if 'مفك عادي' in item or 'flat' in item:
        return 'screwdriver_flat'
    if 'مفك' in item:
        return 'screwdriver_cross'
    if 'صندوق' in item or 'شنتة' in item:
        return 'toolbox_box'
    if 'فرد شحم' in item or 'grease' in item:
        return 'grease_gun'
    if 'فلاتر' in item or 'بايبات' in item or 'بسطون' in item:
        return 'specialty_wrench'
    if category_id == 'sockets':
        return 'socket'
    if category_id in ['combination_wrenches', 'open_wrenches']:
        return 'wrench'
    if category_id in ['torx_keys', 'hex_keys']:
        return 'allen_key'
    return 'wrench'

def english_name(category_id, raw_item, spec):
    raw = str(raw_item).strip()
    translations = {
        'مفك عادي كبير': 'Flat Screwdriver (Large)',
        'مفك عادي وسط': 'Flat Screwdriver (Medium)',
        'مفك عادي وسط صغير': 'Flat Screwdriver (Medium-Small)',
        'مفك عادي صغير': 'Flat Screwdriver (Small)',
        'مفك عادي قصير': 'Flat Screwdriver (Stubby / Short)',
        'مفك عادي طويل': 'Flat Screwdriver (Long)',
        'مفك مصلب كبير': 'Phillips Screwdriver (Large)',
        'مفك مصلب وسط': 'Phillips Screwdriver (Medium)',
        'مفك مصلب صغير': 'Phillips Screwdriver (Small)',
        'مفك مصلب قصير': 'Phillips Screwdriver (Stubby / Short)',
        'مفك مصلب رفيع': 'Phillips Screwdriver (Slim / Precision)',
        'مفك مصلب وسط (امريكي)': 'Phillips Screwdriver Medium (US Spec)',
        'مفك مصلب كبير راس حديد': 'Heavy-Duty Impact Phillips Screwdriver (Large)',
        'يدة 1/2 اوتوماتك': '1/2" Ratchet Handle (Quick-Release)',
        'يدة 1/2 اوتوماتك torx': '1/2" Torx Ratchet Handle',
        'يدة 1/2 اوتوماتك stanly': 'Stanley 1/2" Ratchet Handle',
        'يدة 1/2 اوتوماتك STALLION': 'Stallion 1/2" Ratchet Handle',
        'يدة 1/2 اوتوماتك MITO': 'Mitoloy 1/2" Ratchet Handle',
        'يدة 1/2': '1/2" Drive Handle',
        'يدة 1/2 stanly': 'Stanley 1/2" Drive Handle',
        'يدة 1/2 torx': '1/2" Torx Drive Handle',
        'يدة 1/2  متحركة  torx': '1/2" Swivel Flex Handle Torx',
        'يدة عادية 1/2': '1/2" Standard Breaker Bar',
        'يدة عادية 3/4': '3/4" Standard Breaker Bar',
        'يدة 3/4': '3/4" Heavy Duty Drive Handle',
        'يدة عاديةL 3/4': '3/4" L-Handle Breaker Bar',
        'يدة (1/2+)-(1/2+) L': '1/2" L-Handle Drive',
        'وصلة طويلة 1/2': '1/2" Extension Bar (Long)',
        'وصلة طويلة 1/2  torx': '1/2" Torx Extension Bar (Long)',
        'وصلة طويلة 1/2  stanly': 'Stanley 1/2" Extension Bar (Long)',
        'وصلة وسط 1/2': '1/2" Extension Bar (Medium)',
        'وصلة وسط 1/2 torx': '1/2" Torx Extension Bar (Medium)',
        'وصلة قصيرة 1/2': '1/2" Extension Bar (Short)',
        'وصلة قصيرة 1/2 torx': '1/2" Torx Extension Bar (Short)',
        'وصلة قصيرة 1/2  stanly': 'Stanley 1/2" Extension Bar (Short)',
        'وصلة لعاب 1/2': '1/2" Universal Swivel Joint',
        'وصلة لعاب 1/2  torx': '1/2" Torx Universal Swivel Joint',
        'وصلة لعاب 1/2  stanly': 'Stanley 1/2" Universal Swivel Joint',
        'وصلة لعاب 1/2  MOTO': 'Mitoloy 1/2" Universal Swivel Joint',
        'وصلة وسط 3/4': '3/4" Extension Bar (Medium)',
        'جهاز Multi Meter دينار3': 'Digital Multimeter (Pro Series)',
        'جهاز Multi Meter': 'Digital Multimeter (Digital CAT III)',
        'متر قياس 5 متر SHANG': '5M Heavy-Duty Measuring Tape (Shang)',
        'مصباح كشاف': 'High-Intensity Inspection Worklight',
        'مغناطيس لمسك القطع': 'Telescopic Magnetic Pickup Tool',
        'زرادية': 'Standard Combination Pliers',
        'زرادية بوز': 'Long-Nose Needle Pliers',
        'قطاعة': 'Diagonal Cutting Pliers',
        'قطاعة/تعرية اسلاك': 'Wire Stripper & Cable Cutter',
        'بيلاروس': 'Pincer / End Cutting Pliers',
        'بنسة قفال': 'Locking Grip Pliers (Mole Wrench)',
        'بنسة كبس عادية': 'Slip-Joint / Crimping Pliers',
        'بنسة كبس بوز': 'Needle-Nose Locking Pliers',
        'بنسة عادية بوز': 'Long Needle-Nose Pliers',
        'بنسة وسط عادية': 'Medium Utility Pliers',
        'بنسة سناب رنج': 'Snap Ring Circlip Pliers',
        'كماشة /بنسة وسط': 'Universal Medium Pincers',
        'مطرقة حديد': 'Steel Machinist Hammer (Machinist 500g)',
        'مطرقة بلاستيك': 'Non-Marring Rubber/Nylon Soft Mallet',
        'شاكوش خشب صغير': 'Wooden Handle Precision Hammer',
        'مبرد حديد': 'Flat Bastard Steel File 10"',
        'منشار حديد': 'Junior/Standard Hacksaw 12"',
        'مشحاف وسط': 'Medium Scraper / Putty Knife',
        'مشحاف وسط ': 'Medium Scraper / Putty Knife',
        'كتر': 'Heavy-Duty Retractable Utility Knife',
        'فرد شحم': 'Manual Lever Grease Gun 400cc',
        'مفتاح بسطون': 'Piston Brake Tool / Caliper Compressor',
        'مفتاح بايبات اخضر (pipe wrench)': 'Heavy-Duty Pipe Wrench 14"',
        'مفتاح فلاتر قماش selpro': 'Selpro Strap Oil Filter Wrench',
        'شنتة عدا': 'Heavy-Duty Ballistic Tool Bag',
        'صندوق عدة حديد': '3-Tier Cantilever Heavy Steel Toolbox',
        'صندوق عدة BIG RED': 'Big Red Professional Steel Toolbox',
        'طقم النكيه ستانلي مشرشر': 'Stanley Torx Key Folding Set',
        'طقم بكسات 6 stanly (1/4)': 'Stanley 1/4" Hex Socket Set (6-Piece)',
        'طقم بكسات T طويل': 'Deep T-Handle Socket Set',
        'طقم راس مفك عادي stanly': 'Stanley Screwdriver Bits Set (Flat)',
        'طقم راس مشرشر والنكيه stanly': 'Stanley Torx & Hex Insert Bits Set',
        'طقم شق- رنج(اتوماتك لعاب) Torx': 'Torx Flex-Head Ratcheting Spanners',
        'طقم مبرد صغير 6 حبات': '6-Piece Needle File Set',
        'طقم سناب رنج (ازرق)': '4-Piece Blue Circlip Snap Ring Pliers Set',
        'طقم سناب رنج (اخضر)': '4-Piece Green Circlip Snap Ring Pliers Set',
        'طقم مبرد': 'Mechanic Engineering File Set',
        'للداخل': 'Internal Straight Snap Ring Plier',
        'للخارج': 'External Straight Snap Ring Plier',
        'للداخل زاوية': 'Internal 90° Bent Snap Ring Plier',
        'للخارج زاوية': 'External 90° Bent Snap Ring Plier',
        'مستطيل': 'Flat Hand File',
        'مربع': 'Square File',
        'دائري': 'Round Rat-Tail File',
        'مثلث': 'Triangular Three-Square File',
        'مقطع دائري': 'Precision Round Needle File',
        'مقطع مثلث': 'Precision Triangle Needle File',
        'مقطع مسطح حاد': 'Precision Knife/Flat Needle File',
        'مقطع مسطح': 'Precision Flat Needle File',
        'تحويلة 3/4-1/2 اسود': '3/4" to 1/2" Impact Socket Reducer / Adapter'
    }
    
    if raw in translations:
        return translations[raw]
    for k, v in translations.items():
        if k in raw:
            return v
            
    if category_id == 'sockets':
        return f'Socket {spec or raw}mm'
    if category_id == 'combination_wrenches':
        return f'Combination Spanner {spec or raw}mm'
    if category_id == 'open_wrenches':
        return f'Open-End Spanner {spec or raw}mm'
    if category_id == 'torx_keys':
        return f'Torx Key {spec or raw}'
    if category_id == 'hex_keys':
        return f'Hex Allen Key {spec or raw}mm'
        
    return f'{raw}'

def main():
    files = [
        'Ahmad Jawawdeh.xlsx',
        'mohammad azzam.xlsx',
        'mohammad khalaf alharsa.xlsx',
        'Motazz lutfi.xlsx',
        'Sameer Almoji.xlsx',
        'SAYED ABULLATEEF.xlsx',
        'Sayyed Khairy.xlsx',
        'Yaseen Salamah.xlsx'
    ]
    
    slug_map = {
        'Ahmad Jawawdeh.xlsx': ('ahmad-jawawdeh', 'Ahmad Jawawdeh', 'أحمد جواودة', 'Senior Mechanical Specialist'),
        'mohammad azzam.xlsx': ('mohammad-azzam', 'Mohammad Azzam Al-Mousa', 'محمد عزام الموسى', 'Heavy Equipment Technician'),
        'mohammad khalaf alharsa.xlsx': ('mohammad-khalaf', 'Mohammad Khalaf Al-Harsa', 'محمد خليف الهرسة', 'Hydraulics & Systems Tech'),
        'Motazz lutfi.xlsx': ('motazz-lutfi', 'Motazz Lutfi Abdelaal', 'معتز لطفي عبدالعال', 'Field Operations Specialist'),
        'Sameer Almoji.xlsx': ('sameer-almoji', 'Sameer Al-Moji', 'سمير الموجي', 'Master Diagnostics Technician'),
        'SAYED ABULLATEEF.xlsx': ('sayed-abullateef', 'Sayed Abdullateef Abu Zaid', 'سيد عبداللطيف السيد أبو زيد', 'Powertrain & Mechanical Tech'),
        'Sayyed Khairy.xlsx': ('sayyed-khairy', 'Sayyed Khairy', 'سيد خيري', 'Field Service Technician'),
        'Yaseen Salamah.xlsx': ('yaseen-salamah', 'Yaseen Salamah (Abu Ahmad)', 'يس السلامة (أبو أحمد)', 'Workshop Lead Technician')
    }

    technicians = []
    
    for f in files:
        path = os.path.join(tools_dir, f)
        if not os.path.exists(path):
            continue
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet = wb.active
        
        tech_name_ar = ''
        delivery_date = ''
        for r in range(1, 4):
            for c in range(1, sheet.max_column+1):
                v = sheet.cell(r, c).value
                if v and 'الاسم' in str(v):
                    raw = str(v).strip()
                    if 'تاريخ الاستلام' in raw:
                        parts = raw.split('تاريخ الاستلام')
                        tech_name_ar = parts[0].replace('الاسم', '').replace(':', '').strip()
                        delivery_date = parts[1].replace(':', '').strip()
                    else:
                        tech_name_ar = raw.replace('الاسم', '').replace(':', '').strip()
                    break
            if tech_name_ar:
                break
                
        slug, name_en, name_ar_std, role = slug_map.get(f, (f.replace('.xlsx','').lower(), f.replace('.xlsx',''), tech_name_ar, 'Technician'))
        
        tool_items = []
        item_id = 1
        
        for c in range(1, sheet.max_column + 1):
            current_category = None
            current_qty_col = None
            
            next_c = c + 1
            if next_c <= sheet.max_column:
                for r_test in range(1, 6):
                    val_test = clean_val(sheet.cell(r_test, next_c).value)
                    if val_test in ['كمية', 'الكمية']:
                        current_qty_col = next_c
                        break

            r = 1
            while r <= sheet.max_row:
                cell = sheet.cell(r, c)
                val = clean_val(cell.value)
                
                if val and (
                    any(val.startswith(k) for k in ['طقم', 'شق-', 'شق -', 'شق_', 'النكيه', 'مفتاح شق', 'ادوات اخرى', 'مفكات'])
                    or (val in ['شق-رنج', 'شق-شق', 'النكيه مشرشر', 'النكيه', 'طقم النكيه عادي'])
                ):
                    current_category = val
                    r += 1
                    continue
                    
                if current_category and val:
                    if val in ['مـــفــقــود', 'تــــالــــف', 'لم يتم التسليم', 'كمية', 'الكمية', 'تسلسل']:
                        r += 1
                        continue
                    if 'الاسم :' in val or 'تاريخ الاستلام' in val or 'تاريخ' in val:
                        r += 1
                        continue
                        
                    qty = 1
                    if current_qty_col:
                        q_val = clean_val(sheet.cell(r, current_qty_col).value)
                        if q_val and re.match(r'^\d+(\.\d+)?$', str(q_val)):
                            try:
                                qty = int(float(q_val))
                            except:
                                qty = 1
                                
                    status = get_cell_status(cell)
                    if status == 'good' and current_qty_col:
                        q_status = get_cell_status(sheet.cell(r, current_qty_col))
                        if q_status != 'good':
                            status = q_status
                            
                    cat_id, cat_ar, cat_en = categorize(current_category, val)
                    spec = str(val) if any(ch.isdigit() for ch in str(val)) and len(str(val)) <= 8 else ''
                    name_en_item = english_name(cat_id, val, spec)
                    icon_name = get_icon(cat_id, val)
                    
                    tool_items.append({
                        'id': f'{slug}-tool-{item_id}',
                        'index': item_id,
                        'name': str(val),
                        'nameEn': name_en_item,
                        'category': cat_id,
                        'categoryAr': cat_ar,
                        'categoryEn': cat_en,
                        'specification': spec,
                        'quantity': qty,
                        'status': status,
                        'statusLabelAr': 'سليم / متوفر' if status == 'good' else ('تالف' if status == 'damaged' else ('مفقود' if status == 'missing' else 'لم يتم التسليم')),
                        'statusLabelEn': 'Operational' if status == 'good' else ('Damaged' if status == 'damaged' else ('Missing' if status == 'missing' else 'Pending Delivery')),
                        'icon': icon_name,
                        'row': r,
                        'col': c
                    })
                    item_id += 1
                r += 1
                
        total_items = sum(t['quantity'] for t in tool_items)
        unique_tools = len(tool_items)
        good_count = sum(t['quantity'] for t in tool_items if t['status'] == 'good')
        damaged_count = sum(t['quantity'] for t in tool_items if t['status'] == 'damaged')
        missing_count = sum(t['quantity'] for t in tool_items if t['status'] == 'missing')
        not_delivered_count = sum(t['quantity'] for t in tool_items if t['status'] == 'not_delivered')
        
        categories_dict = {}
        for t in tool_items:
            cid = t['category']
            if cid not in categories_dict:
                categories_dict[cid] = {
                    'id': cid,
                    'titleAr': t['categoryAr'],
                    'titleEn': t['categoryEn'],
                    'count': 0,
                    'totalQty': 0,
                    'damaged': 0,
                    'missing': 0,
                    'pending': 0,
                    'tools': []
                }
            categories_dict[cid]['count'] += 1
            categories_dict[cid]['totalQty'] += t['quantity']
            if t['status'] == 'damaged':
                categories_dict[cid]['damaged'] += t['quantity']
            elif t['status'] == 'missing':
                categories_dict[cid]['missing'] += t['quantity']
            elif t['status'] == 'not_delivered':
                categories_dict[cid]['pending'] += t['quantity']
            categories_dict[cid]['tools'].append(t)
            
        technicians.append({
            'slug': slug,
            'name': name_ar_std,
            'nameEn': name_en,
            'role': role,
            'deliveryDate': delivery_date or 'N/A',
            'file': f,
            'stats': {
                'uniqueTools': unique_tools,
                'totalQuantity': total_items,
                'goodCount': good_count,
                'damagedCount': damaged_count,
                'missingCount': missing_count,
                'notDeliveredCount': not_delivered_count,
                'operationalRate': round((good_count / total_items * 100) if total_items > 0 else 100, 1)
            },
            'categories': list(categories_dict.values()),
            'tools': tool_items
        })
        
    os.makedirs(r'c:\Users\Motasem.ghanem\EQP-System\frontend\data', exist_ok=True)
    out_path = r'c:\Users\Motasem.ghanem\EQP-System\frontend\data\techniciansToolboxes.json'
    with open(out_path, 'w', encoding='utf-8') as out:
        json.dump(technicians, out, ensure_ascii=False, indent=2)

    print(f'Successfully parsed {len(technicians)} technicians and saved to {out_path}')
    for t in technicians:
        print(f"- {t['nameEn']} ({t['name']}): {t['stats']['uniqueTools']} unique, {t['stats']['totalQuantity']} total pcs, {t['stats']['operationalRate']}% readiness")

if __name__ == '__main__':
    main()
