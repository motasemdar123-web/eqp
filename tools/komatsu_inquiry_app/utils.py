import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def parse_input_data(file_or_text):
    """
    Parses an uploaded file (CSV, Excel) or raw text input.
    Returns a DataFrame with standardized columns: ['Part_Number', 'Quantity', 'Original_Row']
    """
    df = None
    if isinstance(file_or_text, str):
        lines = [line.strip() for line in file_or_text.strip().splitlines() if line.strip()]
        rows = []
        for idx, line in enumerate(lines, start=1):
            parts = [p.strip() for p in line.replace('\t', ',').replace(';', ',').split(',') if p.strip()]
            if not parts:
                continue
            part_no = parts[0]
            if part_no.lower() in ['sn', 'part number', 'part no', 'part_number', 'item', 'part', 'number']:
                continue
            qty = 1
            if len(parts) > 1:
                try:
                    qty = int(float(parts[1]))
                except ValueError:
                    qty = 1
            rows.append({'Part_Number': part_no, 'Quantity': qty, 'Original_Row': idx})
        df = pd.DataFrame(rows)
    else:
        filename = getattr(file_or_text, 'name', '').lower()
        if filename.endswith('.csv'):
            try:
                df = pd.read_csv(file_or_text)
            except Exception:
                file_or_text.seek(0)
                df = pd.read_csv(file_or_text, encoding='latin1')
        elif filename.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file_or_text)
        else:
            try:
                df = pd.read_csv(file_or_text)
            except Exception:
                try:
                    file_or_text.seek(0)
                    df = pd.read_excel(file_or_text)
                except Exception:
                    return pd.DataFrame()

    if df is None or df.empty:
        return pd.DataFrame(columns=['Part_Number', 'Quantity', 'Original_Row'])

    cols = {str(c).strip().lower(): c for c in df.columns}
    
    part_col = None
    qty_col = None

    for candidate in ['sn', 'part number', 'part no', 'partno', 'part_number', 'part_no', 'item number', 'item_no', 'number']:
        if candidate in cols:
            part_col = cols[candidate]
            break
    if not part_col:
        part_col = df.columns[0]

    for candidate in ['qty', 'quantity', 'quantities', 'count', 'amount', 'qty required']:
        if candidate in cols:
            qty_col = cols[candidate]
            break

    clean_df = pd.DataFrame()
    clean_df['Part_Number'] = df[part_col].astype(str).str.strip()
    clean_df = clean_df[clean_df['Part_Number'].str.len() > 0]
    clean_df = clean_df[~clean_df['Part_Number'].str.lower().isin(['nan', 'none', 'null', ''])]

    if qty_col and qty_col in df.columns:
        clean_df['Quantity'] = pd.to_numeric(df[qty_col], errors='coerce').fillna(1).astype(int)
    else:
        clean_df['Quantity'] = 1

    clean_df['Original_Row'] = range(1, len(clean_df) + 1)
    return clean_df.reset_index(drop=True)


def chunk_list(items, chunk_size=12):
    """
    Splits a list into chunks of maximum size chunk_size.
    """
    for i in range(0, len(items), chunk_size):
        yield items[i:i + chunk_size]


def generate_styled_excel(df):
    """
    Creates a styled Excel workbook from the results DataFrame with visual hierarchy
    distinguishing Main Parts from Alternative/Interchangeable parts.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Komatsu Parts Inquiry"
    ws.views.sheetView[0].showGridLines = True

    # Palette
    navy_header = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    no_fill = PatternFill(fill_type=None)  # Clean, no background color for all data rows
    
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    main_font = Font(name="Calibri", size=10, bold=True, color="0F172A")
    alt_font = Font(name="Calibri", size=9.5, color="334155")
    alt_desc_font = Font(name="Calibri", size=9.5, color="475569")

    thin_border = Border(
        left=Side(style='thin', color="E2E8F0"),
        right=Side(style='thin', color="E2E8F0"),
        top=Side(style='thin', color="E2E8F0"),
        bottom=Side(style='thin', color="E2E8F0")
    )

    # Write Headers
    headers = list(df.columns)
    ws.append(headers)

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = navy_header
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[1].height = 28

    # Identify Item Type column index
    type_col_idx = None
    for idx, h in enumerate(headers):
        if 'type' in str(h).lower() or 'alternate' in str(h).lower() or 'item' in str(h).lower():
            type_col_idx = idx
            break

    # Write Data Rows
    for row_idx, row_data in enumerate(df.values, start=2):
        ws.append(list(row_data))
        
        # Check if row is Main or Alternate
        is_alt = False
        if type_col_idx is not None and type_col_idx < len(row_data):
            val_str = str(row_data[type_col_idx]).lower()
            is_alt = ('alt' in val_str or 'yes' in val_str or '↳' in val_str)
        else:
            part_str = str(row_data[0] if len(row_data) > 0 else '')
            is_alt = '↳' in part_str

        ws.row_dimensions[row_idx].height = 20

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            header_name = str(headers[col_idx - 1]).lower()

            cell.fill = no_fill  # No background coloring
            cell.font = alt_desc_font if ('desc' in header_name and is_alt) else (alt_font if is_alt else main_font)
            cell.border = thin_border

            # Alignments
            if any(k in header_name for k in ['price', 'weight', 'mor']):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif any(k in header_name for k in ['stock', 'qty', 'quantity', 'eor', 'order', 'cc', 'ic', 'rank', 'code', 'lt', 'lead']):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif 'part number' in header_name:
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=2 if is_alt else 0)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 13), 48)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
